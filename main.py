#!/usr/bin/env python
# -*- coding: utf-8 -*-
'''
import numpy as np
np.random.seed(4)
import pandas as pd
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from matplotlib.figure import Figure
from sklearn.preprocessing import MinMaxScaler
from keras.models import Sequential
from keras.layers import Dense, LSTM
'''
import pandas as pd
import numpy as np
import matplotlib.pylab as plt

plt.rcParams['figure.figsize'] = (16, 9)
plt.style.use('fast')

from keras.models import Sequential
from keras.layers import Dense,Activation,Flatten
from sklearn.preprocessing import MinMaxScaler

from keras.layers import Input, Embedding, Dense, Flatten, Dropout, concatenate, LSTM
from keras.layers import BatchNormalization, SpatialDropout1D
from keras.callbacks import Callback
from keras.models import Model
from keras.optimizers import Adam


import tkinter as tk
from tkinter.filedialog import askopenfilename
from tkinter.messagebox import showerror
from tkinter import ttk
from tkinter import StringVar, Label, Button
from tkcalendar import Calendar, DateEntry

import time
import datetime
from datetime import date, timedelta
import matplotlib.pyplot as plt
import openpyxl
import xlsxwriter
import os.path
from os import path

#import tkk

class Application(tk.Frame):
    def __init__(self, parent, *args, **kwargs):
        tk.Frame.__init__(self, parent, *args, **kwargs)
        self.parent = parent
        self.fileSelect = ""
        self.parent.title("Predicción de Ventas")
        self.menubar = tk.Menu(parent)
        self.filemenu = tk.Menu(self.menubar, tearoff = 0)
        self.filemenu.add_command(label = "Cargar data", command = self.selectFile)

        self.filemenu.add_separator()
        self.filemenu.add_command(label = "Guardar reporte", command = self.saveReport)
        self.filemenu.add_separator()
        self.filemenu.add_command(label = "Guardar model", command = self.saveModel)
        self.filemenu.add_separator()
        self.filemenu.add_command(label = "Salir", command = root.quit)
        self.menubar.add_cascade(label = "Dataset", menu = self.filemenu)
        self.editmenu = tk.Menu(self.menubar, tearoff=0)

        self.editmenu.add_command(label = "Entrena Red Neuronal", command = self.entrenamiento)

        self.menubar.add_cascade(label = "Entrenamiento", menu = self.editmenu)

        self.helpmenu = tk.Menu(self.menubar, tearoff=0)
        self.helpmenu.add_command(label = "Acerca de")
        self.menubar.add_cascade(label = "Ayuda", menu = self.helpmenu)

        self.parent.config(menu = self.menubar)
        
        self.text1 = StringVar()
        self.text1.set('Cargar archivo .csv')
        
        self.texto1 = ttk.Label(self.parent, textvariable=self.text1)
        self.texto1.grid(row=0)

        self.labelR = Label(self.parent,text="Ingrese rango de fechas: ")
        self.labelR.grid(row=1, column=0)

        self.cal1 = DateEntry(self.parent,dateformat=3,width=12, background='darkblue',foreground='white', borderwidth=4,year =2020,month=2,day=15)
        self.cal1.grid(row=1, column=1)

        self.cal2 = DateEntry(self.parent,dateformat=3,width=12, background='darkblue',foreground='white', borderwidth=4,yeaar =2020,month=2,day=29)
        self.cal2.grid(row=1, column=2)

        self.butInter = Button(self.parent, text ="Predecir", command = self.graficar_predicciones)
        self.butInter.grid(row=1, column=3)

        self.lf = ttk.Labelframe(self.parent, text='Ventas')
        self.lf.grid(row=2, column=0, sticky='nwes', padx=3, pady=3)

        '''
        t = np.arange(0.0,3.0,0.01)
        df = pd.DataFrame({'t':t, 's':((2*t/3)*np.pi+120)})

        fig = Figure(figsize=(5,4), dpi=100)
        ax = fig.add_subplot(111)

        df.plot(x='t', y='s', ax=ax)

        self.canvas = FigureCanvasTkAgg(fig, master=lf)
        self.canvas.draw()
        self.canvas.get_tk_widget().grid(row=0, column=0)
        '''
    
    
    def saveReport(self):
        if path.exists('report.png'):
            workbook = xlsxwriter.Workbook('reporte.xlsx')
            worksheet = workbook.add_worksheet()
            workbook.close()
            wb = openpyxl.load_workbook('reporte.xlsx')
            ws = wb.active

            img = openpyxl.drawing.image.Image('report.png')
            #img.anchor(ws.cell('A1'))

            ws.add_image(img)
            wb.save('reporte.xlsx')


    def saveModel(self):
        print("GUARDADO")

    def graficar_predicciones(self):
        
        
        #fig = Figure(figsize=(5,4), dpi=100)
        '''
        fig = plt.figure(figsize=(6, 5))
        plt.plot(real[0:len(prediccion)],color='red', label='Valor real')
        plt.plot(prediccion, color='blue', label='Predicción')
        plt.ylim(1.1 * np.min(prediccion)/2, 1.1 * np.max(prediccion))
        plt.xlabel('Tiempo')
        plt.ylabel('Ingresos')
        plt.legend()
        plt.savefig("report.png", dpi = 150)
        #plt.show()
        canvas = FigureCanvasTkAgg(fig, master=self.lf)
        canvas.draw()
        canvas.get_tk_widget().grid(row=2, column=2)
        '''
        ##INTERPRETACION
        PASOS = 7
        fechaIni = str(self.cal1.get_date() - datetime.timedelta(days=31))
        fechaFin = str(self.cal1.get_date())
        #fechaFin = str(self.cal2.get_date())
        print(fechaFin)
        print(fechaIni)
        ultimosDias = self.df[fechaIni:fechaFin]
        values = ultimosDias['unidades'].values

        # ensure all data is float
        values = values.astype('float32')
        # normalize features
        scaler = MinMaxScaler(feature_range=(-1, 1))

        values=values.reshape(-1, 1) # esto lo hacemos porque tenemos 1 sola dimension

        scaled = scaler.fit_transform(values)

        reframed = self.series_to_supervised(scaled, PASOS, 1)
        reframed.reset_index(inplace=True, drop=True)

        contador=0
        reframed['weekday']=ultimosDias['weekday']
        reframed['month']=ultimosDias['month']

        for i in range(reframed.index[0],reframed.index[-1]):
            reframed['weekday'].loc[contador]=ultimosDias['weekday'][i+8]
            reframed['month'].loc[contador]=ultimosDias['month'][i+8]
            contador=contador+1
        reframed.head()
        
        reordenado=reframed[ ['weekday','month','var1(t-7)','var1(t-6)','var1(t-5)','var1(t-4)','var1(t-3)','var1(t-2)','var1(t-1)'] ]
        reordenado.dropna(inplace=True)
        values = reordenado.values
        x_test = values[5:, :]
        x_test = x_test.reshape((x_test.shape[0], 1, x_test.shape[1]))
        

        ultDiaSemana = reordenado.weekday[len(reordenado)-1]

        def agregarNuevoValor(x_test,nuevoValor,ultDiaSemana):
            for i in range(x_test.shape[2]-3):
                x_test[0][0][i+2] = x_test[0][0][i+3]
            ultDiaSemana=ultDiaSemana+1
            if ultDiaSemana>6:
                ultDiaSemana=0
            x_test[0][0][0]=ultDiaSemana
            x_test[0][0][1]=12
            x_test[0][0][x_test.shape[2]-1]=nuevoValor
            return x_test,ultDiaSemana

        results=[]
        for i in range(7):
            dia=np.array([x_test[0][0][0]])
            mes=np.array([x_test[0][0][1]])
            valores=np.array([x_test[0][0][2:9]])
            parcial=self.model.predict([dia, mes, valores])
            results.append(parcial[0])
            print('pred',i,x_test)
            x_test,ultDiaSemana=agregarNuevoValor(x_test,parcial[0],ultDiaSemana)

        adimen = [x for x in results]
        inverted = scaler.inverse_transform(adimen)
        #inverted


        prediccionProxSemanaDiciembre = pd.DataFrame(inverted)
        prediccionProxSemanaDiciembre.columns = ['pronostico']
        prediccionProxSemanaDiciembre.plot()
        prediccionProxSemanaDiciembre.to_csv('pronostico.csv')

        print(prediccionProxSemanaDiciembre)

    def selectFile(self):
        fname = askopenfilename(filetypes=(("Archivo Dataset", "*.csv"),
                                           ("Todos los archivos", "*.*") ))
        if fname:
            try:
                self.text1.set(fname)
                self.fileSelect = fname
            except:
                print("")
    
    def series_to_supervised(self,data, n_in=1, n_out=1, dropnan=True):
        n_vars = 1 if type(data) is list else data.shape[1]
        df = pd.DataFrame(data)
        cols, names = list(), list()
        # input sequence (t-n, ... t-1)
        for i in range(n_in, 0, -1):
            cols.append(df.shift(i))
            names += [('var%d(t-%d)' % (j+1, i)) for j in range(n_vars)]
        # forecast sequence (t, t+1, ... t+n)
        for i in range(0, n_out):
            cols.append(df.shift(-i))
            if i == 0:
                names += [('var%d(t)' % (j+1)) for j in range(n_vars)]
            else:
                names += [('var%d(t+%d)' % (j+1, i)) for j in range(n_vars)]
        # put it all together
        agg = pd.concat(cols, axis=1)
        agg.columns = names
        # drop rows with NaN values
        if dropnan:
            agg.dropna(inplace=True)
        return agg

    def crear_modeloFF(self):
        PASOS=7
        model = Sequential() 
        model.add(Dense(PASOS, input_shape=(1,PASOS),activation='tanh'))
        model.add(Flatten())
        model.add(Dense(1, activation='tanh'))
        model.compile(loss='mean_absolute_error',optimizer='Adam',metrics=["mse"])
        model.summary()
        return model
    
    def agregarNuevoValor(self,x_test,nuevoValor):
        for i in range(x_test.shape[2]-3):
            x_test[0][0][i+2] = x_test[0][0][i+3]
        ultDiaSemana=ultDiaSemana+1
        if ultDiaSemana>6:
            ultDiaSemana=0
        x_test[0][0][0]=ultDiaSemana
        x_test[0][0][1]=12
        x_test[0][0][x_test.shape[2]-1]=nuevoValor
        return x_test,ultDiaSemana
    
    def crear_modeloEmbeddings(self):
        PASOS=7
        emb_dias = 2 #tamanio profundidad de embeddings
        emb_meses = 4

        in_dias = Input(shape=[1], name = 'dias')
        emb_dias = Embedding(7+1, emb_dias)(in_dias)
        in_meses = Input(shape=[1], name = 'meses')
        emb_meses = Embedding(12+1, emb_meses)(in_meses)

        in_cli = Input(shape=[PASOS], name = 'cli')

        fe = concatenate([(emb_dias), (emb_meses)])

        x = Flatten()(fe)
        x = Dense(PASOS,activation='tanh')(x)
        outp = Dense(1,activation='tanh')(x)
        model = Model(inputs=[in_dias,in_meses,in_cli], outputs=outp)

        model.compile(loss='mean_absolute_error', 
                    optimizer='adam',
                    metrics=['MSE'])

        model.summary()
        return model

    def entrenamiento(self):
        if "csv" in self.fileSelect:
            '''
            dias = (pd.to_datetime(fechaFin) - pd.to_datetime(fechaIni)).days
            diasValidos = int(20*dias/100)
            finValido = datetime.datetime.strptime(fechaFin, '%Y-%m-%d') 
            inicioValido = finValido - timedelta(days=diasValidos)
            inicioEntrenamiento = datetime.datetime.strptime(fechaIni, '%Y-%m-%d') 
            finEntrenamiento = inicioValido - timedelta(days=1)
            '''
            self.df = pd.read_csv(self.fileSelect,  parse_dates=[0], header=None,index_col=0, names=['fecha','unidades'])
            self.df.head()

            self.df['weekday']=[x.weekday() for x in self.df.index]
            self.df['month']=[x.month for x in self.df.index]
            self.df.head()
            self.df.describe()


            print(self.df.head())
            PASOS=7
            # load dataset
            values = self.df['unidades'].values

            # ensure all data is float
            values = values.astype('float32')
            # normalize features
            scaler = MinMaxScaler(feature_range=(-1, 1))

            values=values.reshape(-1, 1) # esto lo hacemos porque tenemos 1 sola dimension

            scaled = scaler.fit_transform(values)

            reframed = self.series_to_supervised(scaled, PASOS, 1)
            reframed.reset_index(inplace=True, drop=True)

            contador=0
            reframed['weekday']=self.df['weekday']
            reframed['month']=self.df['month']

            for i in range(reframed.index[0],reframed.index[-1]):
                reframed['weekday'].loc[contador]=self.df['weekday'][i+8]
                reframed['month'].loc[contador]=self.df['month'][i+8]
                contador=contador+1
            reframed.head()

            reordenado=reframed[ ['weekday','month','var1(t-7)','var1(t-6)','var1(t-5)','var1(t-4)','var1(t-3)','var1(t-2)','var1(t-1)','var1(t)'] ]
            reordenado.dropna(inplace=True)

            training_data = reordenado.drop('var1(t)',axis=1)#.values
            target_data=reordenado['var1(t)']
            #training_data.head()
            valid_data = training_data[len(values)-30:len(values)]
            valid_target=target_data[len(values)-30:len(values)]

            training_data = training_data[0:len(values)-30]
            target_data=target_data[0:len(values)-30]
            print(training_data.shape,target_data.shape,valid_data.shape,valid_target.shape)


            EPOCHS=80

            self.model = self.crear_modeloEmbeddings()

            continuas=training_data[['var1(t-7)','var1(t-6)','var1(t-5)','var1(t-4)','var1(t-3)','var1(t-2)','var1(t-1)']]
            valid_continuas=valid_data[['var1(t-7)','var1(t-6)','var1(t-5)','var1(t-4)','var1(t-3)','var1(t-2)','var1(t-1)']]

            history=self.model.fit([training_data['weekday'],training_data['month'],continuas], target_data, epochs=EPOCHS,validation_data=([valid_data['weekday'],valid_data['month'],valid_continuas],valid_target))
            '''
            plt.scatter(range(len(y_val)),y_val,c='g')
            plt.scatter(range(len(results)),results,c='r')
            plt.title('validate')
            plt.show()

            plt.plot(history.history['loss'])
            plt.title('loss')
            plt.plot(history.history['val_loss'])
            plt.title('validate loss')
            plt.show()
            '''

            results=self.model.predict([valid_data['weekday'],valid_data['month'],valid_continuas])

            plt.scatter(range(len(valid_target)),valid_target,c='g')
            plt.scatter(range(len(results)),results,c='r')
            plt.title('validate')
            plt.show()

            plt.plot(history.history['loss'], label='loss')
            plt.title('loss')
            plt.plot(history.history['val_loss'], label='val loss')
            plt.title('validate loss')
            plt.legend(loc='best')
            plt.show()

            '''
            plt.title('Accuracy')
            plt.plot(history.history['mean_squared_error'])
            plt.show()
            '''


            compara = pd.DataFrame(np.array([valid_target, [x[0] for x in results]])).transpose()
            compara.columns = ['real', 'prediccion']

            inverted = scaler.inverse_transform(compara.values)

            compara2 = pd.DataFrame(inverted)
            compara2.columns = ['real', 'prediccion']
            compara2['diferencia'] = compara2['real'] - compara2['prediccion']
            
            compara2['real'].plot()
            compara2['prediccion'].plot()



            




            '''

            dataset = pd.read_csv(self.fileSelect, index_col='Date', parse_dates=['Date'])
            dataset.head()
            set_entrenamiento = dataset[inicioEntrenamiento.strftime("%Y-%m-%d"):finEntrenamiento.strftime("%Y-%m-%d")].iloc[:,0:1]
            set_validacion = dataset[inicioValido.strftime("%Y-%m-%d"):finValido.strftime("%Y-%m-%d")].iloc[:,0:1]  #20%

            set_entrenamiento['Cantidad'].plot(legend=True)
            set_validacion['Cantidad'].plot(legend=True)

            fig = plt.figure(figsize=(6, 5))
            plt.plot(set_entrenamiento['Cantidad'])
            plt.plot(set_validacion['Cantidad'])
            plt.xticks(rotation='vertical')
            plt.subplots_adjust(bottom=.3)
            plt.legend(['Entrenamiento ('+inicioEntrenamiento.strftime("%Y-%m")+' a '+finEntrenamiento.strftime("%Y-%m")+')', 'Validación ('+finValido.strftime("%Y-%m")+')'])
            canvas = FigureCanvasTkAgg(fig, master=self.lf)
            canvas.draw()
            canvas.get_tk_widget().grid(row=2, column=1)

            sc = MinMaxScaler(feature_range=(0,1))
            set_entrenamiento_escalado = sc.fit_transform(set_entrenamiento)

            time_step = 60
            X_train = []
            Y_train = []
            m = len(set_entrenamiento_escalado)

            for i in range(time_step,m):
                X_train.append(set_entrenamiento_escalado[i-time_step:i,0])
                Y_train.append(set_entrenamiento_escalado[i,0])
            X_train, Y_train = np.array(X_train), np.array(Y_train)
            X_train = np.reshape(X_train, (X_train.shape[0], X_train.shape[1], 1))

            dim_entrada = (X_train.shape[1],1)
            dim_salida = 1
            na = 50

            modelo = Sequential()
            modelo.add(LSTM(units=na, input_shape=dim_entrada))
            modelo.add(Dense(units=dim_salida))
            modelo.compile(optimizer='rmsprop', loss='mse')
            modelo.fit(X_train,Y_train,epochs=600,batch_size=32)

            x_test = set_validacion.values
            x_test = sc.transform(x_test)

            X_test = []
            for i in range(time_step,len(x_test)):
                X_test.append(x_test[i-time_step:i,0])
            X_test = np.array(X_test)
            X_test = np.reshape(X_test, (X_test.shape[0],X_test.shape[1],1))

            prediccion = modelo.predict(X_test)
            prediccion = sc.inverse_transform(prediccion)

            self.graficar_predicciones(set_validacion.values,prediccion)
            '''



if __name__ == "__main__":
    root = tk.Tk()
    w, h = root.winfo_screenwidth(), root.winfo_screenheight()
    root.wm_state('zoomed')
    Application(root)#.pack(side="top", fill="both", expand=True)
    root.mainloop()
