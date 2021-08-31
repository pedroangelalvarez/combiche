import openpyxl
from pathlib import Path

xlsx_file = Path('./', 'datacombiche.xlsx')
wb_obj = openpyxl.load_workbook(xlsx_file)
sheet = wb_obj.active

col_names = []
count = 0
'''
for column in sheet.iter_cols(1,2397):
    col_names.append(str(column[count].value).split(",")[0])
    count += 1
'''
valores = {}
for i in range(1,13035):
    #print(sheet.cell(row=i, column=1).value)
    cad = str(sheet.cell(row=i, column=1).value)[:10]
    #print(cad)
    if not cad in valores:
        valores[str(cad)]=0
    valores[str(cad)]+=1
    
#print(valores)
f = open("datacombichee.txt", "a")


for key in valores.keys():
    print(key+"\t"+str(valores[key]))
    f.write(key+"\t"+str(valores[key])+"\n")
f.close()